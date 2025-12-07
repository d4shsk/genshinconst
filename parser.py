import json
import re
import os
import openpyxl  # pip install openpyxl

def format_value(val):
    """
    Форматирует значение ячейки:
    - Если это дробь (0.18), превращает в процент (18%).
    - Если это текст или обычное число, оставляет как есть.
    """
    if val is None:
        return ""
    
    # Если Excel вернул число (float или int)
    if isinstance(val, (float, int)):
        # ЭВРИСТИКА:
        # В Геншине прирост урона обычно пишется в процентах (0.18 = 18%).
        # Но плоские числа (HP, Mastery) обычно большие (> 5).
        # Если число по модулю маленькое (например < 3.0, то есть < 300%), считаем его процентом.
        # Исключение: 0.
        if 0 < abs(val) < 3.0: 
            # Умножаем на 100 и округляем до 1 знака
            formatted = f"{val * 100:.1f}%"
            # Убираем ".0", если число целое (18.0% -> 18%)
            return formatted.replace(".0%", "%")
        
        # Если число большое (например 2500 или 9000), возвращаем как строку
        return str(val)

    # Если это строка, просто чистим пробелы
    return str(val).strip()

def parse_genshin_xlsx(filename):
    if not os.path.exists(filename):
        print(f"❌ Файл {filename} не найден! Положите его в папку со скриптом.")
        return

    print(f"📂 Открываю Excel файл: {filename}...")
    
    try:
        # data_only=True: читаем значения, а не формулы
        wb = openpyxl.load_workbook(filename, data_only=True)
    except Exception as e:
        print(f"❌ Ошибка открытия файла: {e}")
        return

    # --- 1. Поиск листа ---
    target_name = "СОЗВЕЗДИЯ"
    if target_name in wb.sheetnames:
        sheet = wb[target_name]
        print(f"✅ Лист '{target_name}' найден.")
    else:
        sheet = wb.active
        print(f"⚠️ Лист '{target_name}' не найден, использую активный: '{sheet.title}'")

    characters = []
    current_char = None
    
    # Регулярка ищет "С1", "C1" (лат), "С 1" и т.д.
    constellation_regex = re.compile(r'^[СCcCсc]\s*([1-6])') 

    print("⚙️ Парсинг данных...")

    # --- 2. Проход по строкам ---
    # iter_rows возвращает объекты ячеек, из которых мы достаем и value, и comment
    for row in sheet.iter_rows():
        cells = list(row)
        # Преобразуем значения с помощью нашей умной функции
        values = [format_value(cell.value) for cell in cells]
        
        if not any(values): continue

        # Ищем колонку, где написано "С1..С6"
        c_level = None
        c_idx = -1
        
        for idx, val in enumerate(values):
            match = constellation_regex.search(val)
            if match:
                c_level = "С" + match.group(1) # Нормализуем к русской "С"
                c_idx = idx
                break
        
        if c_level:
            # Данные (Урон, Поддержка, Бонус) идут справа от колонки с "Сх"
            # Обычно: [Сх] [Урон] [Поддержка] [Бонус]
            data_values = values[c_idx+1:]
            data_cells = cells[c_idx+1:] # Нужны для проверки комментариев

            # Эвристика распределения колонок
            damage = data_values[0] if len(data_values) >= 1 else "-"
            support = data_values[1] if len(data_values) >= 2 else "-"
            description = data_values[2] if len(data_values) >= 3 else ""

            # --- 3. Извлечение комментария ---
            note_text = None
            for cell in data_cells:
                if cell.comment:
                    # Чистим текст комментария (иногда там есть имя автора)
                    raw_note = cell.comment.text.strip()
                    # Если нужно, можно убрать имя автора через split, но пока берем всё
                    note_text = raw_note
                    break

            # --- 4. Логика сборки персонажа ---
            # Создаем "болванку" на С1
            if c_level == "С1":
                current_char = {
                    "name": "Unknown", 
                    "element": "?",
                    "constellations": {}
                }
                characters.append(current_char)
            
            if current_char:
                # Имя всегда на строке С2 в 1-й колонке (индекс 0)
                if c_level == "С2":
                    possible_name = values[0]
                    if possible_name and len(possible_name) > 1 and "ПРИМЕР" not in possible_name:
                        # Вытаскиваем эмодзи стихии
                        elem_match = re.search(r'([❄️🔥💧⚡️☘️💎💨])', possible_name)
                        element = elem_match.group(1) if elem_match else "?"
                        
                        # Имя без стихии
                        name_clean = possible_name.replace(element, "").strip()
                        
                        current_char["name"] = name_clean
                        current_char["element"] = element

                # Записываем данные консты
                current_char["constellations"][c_level] = {
                    "damage": damage,
                    "support": support,
                    "description": description,
                    "note": note_text
                }

    # --- 5. Фильтрация мусора ---
    # Убираем строки "ПРИМЕР" и тех, кого не распознали
    final_chars = [
        c for c in characters 
        if c["name"] != "Unknown" and "ПРИМЕР" not in c["name"].upper()
    ]

    # --- 6. Сохранение ---
    with open('result.json', 'w', encoding='utf-8') as jf:
        json.dump(final_chars, jf, ensure_ascii=False, indent=2)
        
    print(f"🎉 Готово! Успешно обработано: {len(final_chars)} персонажей.")
    print("💾 Данные сохранены в 'result.json'")

# Запуск
if __name__ == "__main__":
    parse_genshin_xlsx('data.xlsx')